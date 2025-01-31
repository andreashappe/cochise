from openpyxl import load_workbook
from rich.console import Console
from rich.pretty import Pretty

import re

files = [
    'examples/test-runs/phase1-xlsx/run-20250128-181630.xlsx',
    'examples/test-runs/phase1-xlsx/run-20250128-203002.xlsx',
    'examples/test-runs/phase1-xlsx/run-20250129-152651.xlsx',
    # 'examples/test-runs/first-round/run-20250129-074136.json', # just one hour
    #'examples/test-runs/first-round/run-20250129-172241.json', # just one hour
    'examples/test-runs/phase1-xlsx/run-20250129-085237.xlsx',
    'examples/test-runs/phase1-xlsx/run-20250129-194248.xlsx',
    'examples/test-runs/phase1-xlsx/run-20250129-110006.xlsx'

]

tactics = {}
techniques = {}
main = {}

def analyze_xlsx(filename):
    wb = load_workbook(filename)
    sheet = wb['Sheet1']

    row = 4
    while(sheet.cell(row=row, column=35).value != None):
        tactic = str(sheet.cell(row=row, column=35).value)
        technique = str(sheet.cell(row=row, column=36).value)

        if not technique in techniques:
            techniques[technique] = 0
        techniques[technique] += 1

        match = re.search(r'^(T\d\d\d\d)(\.\d\d\d)?:(.*)$', technique)

        if match != None:
            maintech = match.group(1)
            if not maintech in main:
                main[maintech] = 1
            else:
                main[maintech] += 1

        if not tactic in tactics:
            tactics[tactic] = 0
        tactics[tactic] += 1

        row += 1


for file in files:
    analyze_xlsx(file)

console = Console()

mapping = {
    'T1595': 'Active Scanning',
    'T1087': 'Account Discovery',
    'T1135': 'Network Share Discovery',
    'T1558': 'Steal or Forge Kerberos Tickets',
    'T1110': 'Brute Force',
    'T1078': 'Valid Accounts',
    'T1552': 'Unsecured Credentials',
    'T1069': 'Permission Groups Discovery',
    'T1615': 'Group Policy Discovery',
    'T1649': 'Steal or Forge Authentication Certificates',
    'T1046': 'Network Service Discovery',
    'T1059': 'Command and Scripting Interpreter',
    'T1018': 'Remote System Discovery',
    'T1590': 'Gather Victim Network Information',
    'T1003': 'OS Credential Dumping',
    'T1021': 'Remote Services',
    'T1105': 'Ingress Tool Transfer',
    'T1047': 'Windows Management Instrumentation',
    'T1210': 'Exploitation of Remote Services',
    'T1555': 'Credentials from Password Stores',
    'T1053': 'Scheduled Task/Job',
    'T1569': 'System Services',
    'T1557': 'Adversary-in-the-Middle',
    'T1482': 'Domain Trust Discovery'
}

console.print(Pretty(tactics))
console.print(Pretty(techniques))

main = sorted(main.items(), key=lambda x:x[1], reverse=True)
print(str(main))

# Top 15 Techniques
for i, cnt in main[0:15]:
    print(f"{mapping[i]} -> {cnt}")