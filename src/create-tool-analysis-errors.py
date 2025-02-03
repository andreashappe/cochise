
from openpyxl import load_workbook
from rich.console import Console
from rich.pretty import Pretty

import re

files = [
    'examples/test-runs/phase1-xlsx/run-20250128-181630_executed_commands.xlsx',
    'examples/test-runs/phase1-xlsx/run-20250128-203002_executed_commands.xlsx',
    'examples/test-runs/phase1-xlsx/run-20250129-152651_executed_commands.xlsx',
    # 'examples/test-runs/first-round/run-20250129-074136.json', # just one hour
    #'examples/test-runs/first-round/run-20250129-172241.json', # just one hour
    'examples/test-runs/phase1-xlsx/run-20250129-085237_executed_commands.xlsx',
    'examples/test-runs/phase1-xlsx/run-20250129-194248_executed_commands.xlsx',
    'examples/test-runs/phase1-xlsx/run-20250129-110006_executed_commands.xlsx'

]

def get_value(input:str) -> bool:
    if input == '=FALSE()' or input == False:
        return False
    if input == '=TRUE()' or input == True:
        return True
    print(str(input))
    assert(False)

cmds = {}

def incr_cmd_type(cmd, typ):
    if not cmd in cmds.keys():
        cmds[cmd] = {}
    if not typ in cmds[cmd]:
        cmds[cmd][typ] = 1
    else:
        cmds[cmd][typ] += 1

def analyze_xlsx(filename):
    wb = load_workbook(filename)
    sheet = wb['Sheet1']

    invalid_param_syntax = 0
    unknown_command = 0
    invalid_param_semantics = 0
    all_cmds = 0

    row = 2
    while(sheet.cell(row=row, column=1).value != None):

        cmd = str(sheet.cell(row=row, column=1).value)

        incr_cmd_type(cmd, 'count')
        if get_value(sheet.cell(row=row, column=5).value) == True:
            invalid_param_syntax += 1
            incr_cmd_type(cmd, 'syntax')
            incr_cmd_type(cmd, 'error')
        #if get_value(sheet.cell(row=row, column=6).value) == True:
        #    unknown_command += 1
        #    incr_cmd_type(cmd, 'unknown')
        #    incr_cmd_type(cmd, 'error')
        if get_value(sheet.cell(row=row, column=7).value) == True:
            invalid_param_semantics += 1
            incr_cmd_type(cmd, 'semantics')
            incr_cmd_type(cmd, 'error')

        all_cmds += 1
        row += 1

    return {
        'unknown_command': unknown_command,
        'invalid_param_syntax': invalid_param_syntax,
        'invalid_param_semantic': invalid_param_semantics,
        'sum': (unknown_command + invalid_param_semantics + invalid_param_syntax)/all_cmds
    }

for file in files:
    print(f"now doing file {file}")
    print(str(analyze_xlsx(file)))

top15 = [
    'smbclient', 'nxc', 'cat', 'echo', 'nmap', 'rpcclient', 'impacket-GetUserSPNs', 'john', 'impacket-GetNPUsers', 'hashcat', 'impacket-mssqlclient', 'netexec', 'impacket-smbexec', 'impacket-secretsdump', 'impacket-GetADUsers', 'ls'
]

for i in top15:
    if 'syntax' in cmds[i]:
        syntax = cmds[i]['syntax']
    else:
        syntax = 0

    if 'semantics' in cmds[i]:
        semantics = cmds[i]['semantics']
    else:
        semantics = 0

    print(f"{i}: {cmds[i]['error']/cmds[i]['count']} -> {syntax/cmds[i]['count']} / {semantics/cmds[i]['count']} ")

console = Console()